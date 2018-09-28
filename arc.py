import os
import ssl
import urllib
import urllib2
import re
import json
import time
import datetime
import logging
import logging.config
import pyodbc
from email import encoders
from email.header import Header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.utils import parseaddr, formataddr
import smtplib
from openpyxl import Workbook


class ArcModel:
    logging.config.fileConfig("logger.conf")
    logger = None

    def __init__(self, logger_name=None):
        if logger_name:
            self.logger = logging.getLogger(logger_name)
        # self.logger.debug("import arc")
        self._host = 'iar2.arccorp.com'
        self._origin = 'https://iar2.arccorp.com'
        self._accept = 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8'
        self._accept_language = 'zh-CN,zh;q=0.8,en;q=0.6'
        self._accept_encoding = 'gzip, deflate, br'
        self._cache_control = 'max-age=0'
        self._connection = 'keep-alive'
        self._content_type = 'application/x-www-form-urlencoded'
        self._upgrade_insecure_requests = '1'
        self._user_agent = 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.98 Safari/537.36'
        self._cookies = urllib2.HTTPCookieProcessor()
        self._opener = urllib2.build_opener(self._cookies)

    def __save_page(self, page, arcNumber, html):
        # f=open(page+'_'+datetime.datetime.now().strftime('%Y%m%d%H:%M:%S')+'.html','wb')
        path = 'html'
        if not os.path.exists(path):
            os.makedirs(path)

        with open(path + '\\' + page + '_' + arcNumber + '.html', 'wb') as f:
            f.write(html)

    def __save_csv(self, name, is_this_week, fileName, content):
        # print "SAVE: %s ,Is this week: %s, file: %s" % (name, is_this_week, fileName)
        file_path = "day"
        if not is_this_week:
            file_path = "week"

        if name == "mulingpeng" or name == "gttqc02" or name == "gttqc-it":
            name = "all"

        if name == "muling-yww":
            name = "yww"

        if name == "muling-tvo":
            name = "tvo"

        if name == "muling-aca":
            name = "aca"

        path = name + '\\' + file_path
        if not os.path.exists(path):
            os.makedirs(path)

        with open(path + '\\' + fileName, 'wb') as f:
            f.write(content)

    def __try_request(self, req, max_try_num=2):
        res = None
        for tries in range(max_try_num):
            try:
                # self.logger.debug("Request start at %d times" % tries)
                res = self._opener.open(req, timeout=30)
                # self.logger.debug("Request success at %d times" % tries)
                break
            except (urllib2.URLError, ssl.SSLError) as e:
                if tries < (max_try_num - 1):
                    self.logger.info("Request error at %d times, will try again after 2s." % tries)
                    time.sleep(2)
                    continue
                else:
                    self.logger.warning(e)
            except Exception, e:
                self.logger.critical(repr(e))
                break

        # self.logger.debug("TRY REQUEST OVER")
        return res

    def set_certificate_item(self, certificates, error_code, agent_codes):
        certificateItems = []
        if certificates:
            for i in certificates:
                if i[1] not in agent_codes:
                    certificateItems.append(i[1])

        certificate_item_length = len(certificateItems)
        for i in range(0, 3):
            if i >= certificate_item_length:
                certificateItems.append("")

        if not error_code and len(certificateItems) == 3:
            certificateItems.insert(3, "")
        elif error_code and len(certificateItems) == 3:
            certificateItems.insert(0, error_code)

        return certificateItems

    def csv_write(self, file_path, file_name, content):
        if not os.path.exists(file_path):
            os.makedirs(file_path)

        with open(file_path + '\\' + file_name, 'wb') as f:
            f.write(content)

    def login(self, name, password):
        self.logger.debug("START HOME")
        self._opener.open("https://www.arccorp.com", timeout=60)
        self.logger.debug("LOGIN: %s" % name)

        values = {
            'userID': "",
            'user': name,
            'password': password
        }

        url = "https://myarc.arccorp.com/PortalApp/PreLogin.portal"
        data = urllib.urlencode(values)

        headers = {
            'Accept': self._accept,
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': self._accept_language,
            'Cache-Control': self._cache_control,
            'Connection': self._connection,
            'Content-Length': len(data),
            'Content-Type': self._content_type,
            'Host': 'myarc.arccorp.com',
            'Origin': 'https://myarc.arccorp.com',
            'Referer': 'https://myarc.arccorp.com/PortalApp/PreLogin.portal?logout',
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }

        req = urllib2.Request(url, data, headers)
        res = self.__try_request(req)
        if res:
            html = res.read()
            self.__save_page('login', 'login', html)
            return html

    def execute_login(self, name, password, max_try_num=2):
        is_login = False
        for tries in range(max_try_num):
            html = self.login(name, password)
            if not html or (html.find('You are already logged into My ARC') < 0 and html.find('Account Settings:') < 0):
                self.logger.error("LOGIN ERROR: %s, at %d times" % (name, tries))
                time.sleep(1 * 60 * 3)
                continue

            is_login = True
            break

        return is_login

    def iar(self):
        self.logger.debug("GO TO IAR")
        url = "https://iar2.arccorp.com/IAR/"
        res = self.__try_request(url)
        if res:
            html = res.read()
            self.__save_page('iar', 'iar', html)
            return html

        # try:
        #     res = self._opener.open(url)
        #     html = res.read()
        #     self.__save_page('iar', 'iar', html)
        #     return html
        # except urllib2.HTTPError, e:
        #     self.logger.warning(e.code)
        # except urllib2.URLError, e:
        #     self.logger.warning(e.reason)
        # except Exception, e:
        #     self.logger.critical(e)

    def listTransactions(self, ped, action, arcNumber):
        self.logger.debug("GO TO LIST TRANSACTIONS")
        url = "https://iar2.arccorp.com/IAR/listTransactions.do?ped=" + ped + "&action=" + action + "&arcNumber=" + arcNumber
        res = self.__try_request(url)
        if res:
            html = res.read()
            self.__save_page("listTransactions", arcNumber, html)
            return html

    def create_list(self, token, ped, action, arcNumber, viewFromDate, viewToDate, documentNumber, selectedStatusId='',
                    selectedDocumentType='', selectedTransactionType='', selectedFormOfPayment='', dateTypeRadioButtons='ped',
                    selectedNumberOfResults='20', is_next=False, page=0):
        self.logger.debug("CREATE LIST: %s" % arcNumber)
        values = {
            'org.apache.struts.taglib.html.TOKEN': token,
            'arcNumber': arcNumber,
            'ped': ped,
            # 'selectedStatusId': "",
            'selectedStatusId': selectedStatusId,
            'documentNumber': documentNumber,
            'docNumberEnd': documentNumber,
            'selectedDocumentType': selectedDocumentType,
            # 'selectedTransactionType': 'SA',
            'selectedTransactionType': selectedTransactionType,
            # 'selectedFormOfPayment': 'CA',
            'selectedFormOfPayment': selectedFormOfPayment,
            'selectedInternationalIndicator': '',
            'systemProvider': '',
            # 'dateTypeRadioButtons': 'ped',
            'dateTypeRadioButtons': dateTypeRadioButtons,
            # 'viewFromDate': '02APR18',
            'viewFromDate': viewFromDate,
            # 'viewToDate': '08APR18',
            'viewToDate': viewToDate,
            'commTypeRadioButtons': 'commEqualTo',
            'commissionAmount': '',
            'threeDigitCarrierCode': '',
            # 'selectedNumberOfResults': '20',
            'selectedNumberOfResults': selectedNumberOfResults,
            # 'createlistdisabled': "1",
            'list.x': '19',
            'list.y': '8',
            'printOption': '1',
            'printaction': '0'
        }

        if is_next:
            del values['list.x']
            del values['list.y']
            values['next.x'] = "17"
            values['next.y'] = "11"

        url = "https://iar2.arccorp.com/IAR/listTransactions.do"
        data = urllib.urlencode(values)
        headers = {
            'Host': self._host,
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:59.0) Gecko/20100101 Firefox/59.0',
            # 'User-Agent': self._user_agent,
            'Accept': self._accept,
            'Accept-Language': 'zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2',
            # 'Accept-Language': self._accept_language,
            'Accept-Encoding': self._accept_encoding,
            'Referer': 'https://iar2.arccorp.com/IAR/listTransactions.do?ped=' + ped + '&action=' + action + '&arcNumber=' + arcNumber + '',
            'Content-Type': self._content_type,
            'Content-Length': len(data),
            'Connection': self._connection,
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests
        }

        if documentNumber:
            headers['Cache-Control'] = self._cache_control
            headers['Origin'] = self._origin

        req = urllib2.Request(url, data, headers)
        res = self.__try_request(req)
        if res:
            html = res.read()
            page_flag = ""
            if page > 0:
                page_flag = "_P" + str(page)
            self.__save_page("create_list", arcNumber + page_flag, html)
            return html

    def get_csv(self, name, is_this_week, ped, action, arcNumber, token, viewFromDate, viewToDate, dateTypeRadioButtons='ped',
                selectedStatusId='', selectedTransactionType='', selectedFormOfPayment='', tick=''):
        self.logger.debug("DOWNLOAD CSV: " + arcNumber)
        values = {
            'org.apache.struts.taglib.html.TOKEN': token,
            'arcNumber': arcNumber,
            'ped': ped,
            'selectedStatusId': selectedStatusId,
            'documentNumber': '',
            'docNumberEnd': '',
            'selectedDocumentType': '',
            'selectedTransactionType': selectedTransactionType,
            'selectedFormOfPayment': selectedFormOfPayment,
            'selectedInternationalIndicator': '',
            'systemProvider': '',
            'dateTypeRadioButtons': dateTypeRadioButtons,
            'viewFromDate': viewFromDate,
            'viewToDate': viewToDate,
            'commTypeRadioButtons': 'commEqualTo',
            'commissionAmount': '',
            'threeDigitCarrierCode': '',
            'selectedNumberOfResults': '20',
            'printOption': '2',
            'download.x': '35',
            'download.y': '8',
            'action': ''
        }

        if selectedTransactionType:
            del values['action']

            values['download.x'] = "28"
            values['printOption'] = "2"
            values['printaction'] = "0"

        # print values
        url = "https://iar2.arccorp.com/IAR/listTransactions.do"
        data = urllib.urlencode(values)
        headers = {
            'Accept': self._accept,
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': self._accept_language,
            'Cache-Control': self._cache_control,
            'Connection': self._connection,
            # 'Content-Length':'480',
            'Content-Length': len(data),
            'Content-Type': self._content_type,
            'Host': self._host,
            'Origin': self._origin,
            'Referer': 'https://iar2.arccorp.com/IAR/listTransactions.do?ped=' + ped + '&action=' + action + '&arcNumber=' + arcNumber + '',
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }

        if selectedTransactionType:
            headers['Referer'] = "https://iar2.arccorp.com/IAR/listTransactions.do"
        req = urllib2.Request(url, data, headers)
        res = self.__try_request(req)
        if res:
            csv = res.read()
            if csv.find("<link") == -1:
                self.__save_csv(name, is_this_week, arcNumber + ("" if not tick else "_" + tick) + '.csv', csv)
                if selectedTransactionType:
                    # self.logger.debug("ARC: %s, CSV: %s" % (arcNumber, csv))
                    return csv
        else:
            self.logger.warning('Download csv error :' + arcNumber)

    # def search(self, ped, action, arcNumber, token, from_date, to_date, documentNumber, dateTypeRadioButtons='ped'):
    #     self.logger.debug("SEARCH TICKET: %s" % documentNumber)
    #     values = {
    #         'org.apache.struts.taglib.html.TOKEN': token,
    #         'arcNumber': arcNumber,
    #         'ped': ped,
    #         'selectedStatusId': '',
    #         'documentNumber': documentNumber,
    #         'docNumberEnd': documentNumber,
    #         'selectedDocumentType': '',
    #         'selectedTransactionType': '',
    #         'selectedFormOfPayment': '',
    #         'selectedInternationalIndicator': '',
    #         'systemProvider': '',
    #         'dateTypeRadioButtons': dateTypeRadioButtons,
    #         'viewFromDate': from_date,
    #         'viewToDate': to_date,
    #         'commTypeRadioButtons': 'commEqualTo',
    #         'commissionAmount': '',
    #         'threeDigitCarrierCode': '',
    #         'selectedNumberOfResults': '20',
    #         'list.x': '45',
    #         'list.y': '11',
    #         'printOption': '1',
    #         'printaction': '0'
    #     }
    #
    #     url = "https://iar2.arccorp.com/IAR/listTransactions.do"
    #     data = urllib.urlencode(values)
    #
    #     headers = {
    #         'Accept': self._accept,
    #         'Accept-Encoding': 'gzip, deflate, br',
    #         'Accept-Language': self._accept_language,
    #         'Cache-Control': self._cache_control,
    #         'Connection': self._connection,
    #         'Content-Length': len(data),
    #         'Content-Type': self._content_type,
    #         'Host': 'iar2.arccorp.com',
    #         'Origin': 'https://iar2.arccorp.com',
    #         'Referer': 'https://iar2.arccorp.com/IAR/listTransactions.do?ped=' + ped + '&action=' + action + '&arcNumber=' + arcNumber + '',
    #         'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
    #         'User-Agent': self._user_agent
    #     }
    #     req = urllib2.Request(url, data, headers)
    #     res = self.__try_request(req)
    #     if res:
    #         html = res.read()
    #         self.__save_page("search", "search", html)
    #         return html

    def add_old_document(self, token, oldDocumentAirlineCodeFI, oldDocumentNumberFI, seqNum, documentNumber, amountCommission,
                         waiverCode, certificateItems, maskedFC):
        self.logger.debug("ADD OLD DOCUMENT, AIR: %s, DOC NUM: %s, SEQ: %s, TKT: %s." % (oldDocumentAirlineCodeFI,
                          oldDocumentNumberFI, seqNum, documentNumber))

        if waiverCode is None:
            waiverCode = ""

        if maskedFC is None:
            maskedFC = ""

        values = {
            'org.apache.struts.taglib.html.TOKEN': token,
            'amountCommission': amountCommission,
            'maskedFormOfPayment': "CASH",
            'paymentBean.approvalCode': '',
            'paymentBean.extendedPay': "N",
            'miscSupportTypeId': "",
            'waiverCode': waiverCode,
            'certificateItem[0].value': certificateItems[0],
            'certificateItem[1].value': certificateItems[1],
            'certificateItem[2].value': certificateItems[2],
            'certificateItem[3].value': certificateItems[3],
            'error22010': 'false',
            'oldDocumentAirlineCodeFI': oldDocumentAirlineCodeFI,
            'oldDocumentNumberFI': oldDocumentNumberFI,
            'addOldDocumentButton.x': "57",
            'addOldDocumentButton.y': "9",
            'selfSaleIntlTypeId': "",
            'maskedFC': maskedFC
        }

        url = "https://iar2.arccorp.com/IAR/financialDetails.do"
        data = urllib.urlencode(values)

        headers = {
            'Accept': self._accept,
            'Accept-Encoding': self._accept_encoding,
            'Accept-Language': self._accept_language,
            'Connection': self._connection,
            'Content-Length': len(data),
            'Content-Type': self._content_type,
            'Host': self._host,
            'Referer': "https://iar2.arccorp.com/IAR/modifyTran.do?seqNum=" + seqNum + "&documentNumber=" + documentNumber,
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }

        req = urllib2.Request(url, data, headers)
        res = self.__try_request(req)
        if res:
            html = res.read()
            self.__save_page("add_old_document", documentNumber, html)
            return html

    def exchange_input(self, token, documentNumber):
        self.logger.debug("EXCHANGE INPUT.")
        values = {
            'org.apache.struts.taglib.html.TOKEN': token,
            'inputOldDoc.ESAC': "",
            'inputOldDoc.CPUI1234[0].value': "1",
            'inputOldDoc.oldFare': "1",
            'inputOldDoc.oldTotal': "1",
            'inputOldDoc.oldComm': "",
            'inputOldDoc.adminPenalty': "",
            'inputOldDoc.commOnPenalty': "",
            'inputOldDoc.taxCodeList[0].taxCode': "",
            'inputOldDoc.taxCodeList[0].taxAmount': "",
            'inputOldDoc.taxCodeList[1].taxCode': "",
            'inputOldDoc.taxCodeList[1].taxAmount': "",
            'inputOldDoc.taxCodeList[2].taxCode': "",
            'inputOldDoc.taxCodeList[2].taxAmount': "",
            'inputOldDoc.taxCodeList[3].taxCode': "",
            'inputOldDoc.taxCodeList[3].taxAmount': "",
            'inputOldDoc.taxCodeList[4].taxCode': "",
            'inputOldDoc.taxCodeList[4].taxAmount': "",
            'inputOldDoc.taxCodeList[5].taxCode': "",
            'inputOldDoc.taxCodeList[5].taxAmount': "",
            'inputOldDoc.taxCodeList[6].taxCode': "",
            'inputOldDoc.taxCodeList[6].taxAmount': "",
            'inputOldDoc.taxCodeList[7].taxCode': "",
            'inputOldDoc.taxCodeList[7].taxAmount': "",
            'inputOldDoc.taxCodeList[8].taxCode': "",
            'inputOldDoc.taxCodeList[8].taxAmount': "",
            'inputOldDoc.taxCodeList[9].taxCode': "",
            'inputOldDoc.taxCodeList[9].taxAmount': "",
            'inputOldDoc.taxCodeList[10].taxCode': "",
            'inputOldDoc.taxCodeList[10].taxAmount': "",
            'inputOldDoc.taxCodeList[11].taxCode': "",
            'inputOldDoc.taxCodeList[11].taxAmount': "",
            'inputOldDoc.taxCodeList[12].taxCode': "",
            'inputOldDoc.taxCodeList[12].taxAmount': "",
            'inputOldDoc.taxCodeList[13].taxCode': "",
            'inputOldDoc.taxCodeList[13].taxAmount': "",
            'inputOldDoc.taxCodeList[14].taxCode': "",
            'inputOldDoc.taxCodeList[14].taxAmount': "",
            'inputOldDoc.taxCodeList[15].taxCode': "",
            'inputOldDoc.taxCodeList[15].taxAmount': "",
            'inputOldDoc.taxCodeList[16].taxCode': "",
            'inputOldDoc.taxCodeList[16].taxAmount': "",
            'inputOldDoc.taxCodeList[17].taxCode': "",
            'inputOldDoc.taxCodeList[17].taxAmount': "",
            'inputOldDoc.taxCodeList[18].taxCode': "",
            'inputOldDoc.taxCodeList[18].taxAmount': "",
            'inputOldDoc.taxCodeList[19].taxCode': "",
            'inputOldDoc.taxCodeList[19].taxAmount': "",
            'inputOldDoc.taxCodeList[20].taxCode': "",
            'inputOldDoc.taxCodeList[20].taxAmount': "",
            'inputOldDoc.pfcList[0].value': "",
            'inputOldDoc.pfcList[1].value': "",
            'inputOldDoc.pfcList[2].value': "",
            'inputOldDoc.pfcList[3].value': "",
            'addButton.x': "28",
            'addButton.y': "8"
        }

        url = "https://iar2.arccorp.com/IAR/exchangeInput.do"
        data = urllib.urlencode(values)

        headers = {
            'Accept': self._accept,
            'Accept-Encoding': self._accept_encoding,
            'Accept-Language': self._accept_language,
            'Connection': self._connection,
            'Content-Length': len(data),
            'Content-Type': self._content_type,
            'Host': self._host,
            'Referer': "https://iar2.arccorp.com/IAR/financialDetails.do",
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }

        req = urllib2.Request(url, data, headers)
        res = self.__try_request(req)
        if res:
            html = res.read()
            self.__save_page("exchange_input", documentNumber, html)
            return html

    def exchange_summary(self, token, documentNumber):
        self.logger.debug("EXCHANGE SUMMARY.")
        values = {
            'org.apache.struts.taglib.html.TOKEN': token,
            'navButton1.x': "79",
            'navButton1.y': "16",
            'adjustment': "",
            'maskedFormOfPayment': "CA54000000",
            'paymentBeanExchange.approvalCode': "",
            'paymentBeanExchange.extendedPay': "N"
        }

        url = "https://iar2.arccorp.com/IAR/financialDetails.do"
        data = urllib.urlencode(values)

        headers = {
            'Accept': self._accept,
            'Accept-Encoding': self._accept_encoding,
            'Accept-Language': self._accept_language,
            'Connection': self._connection,
            'Content-Length': len(data),
            'Content-Type': self._content_type,
            'Host': self._host,
            'Referer': "https://iar2.arccorp.com/IAR/exchangeInput.do",
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }

        req = urllib2.Request(url, data, headers)
        res = self.__try_request(req)
        if res:
            html = res.read()
            self.__save_page("exchange_summary", documentNumber, html)
            return html

    def remove_old_document(self, token, commission, documentNumber, waiverCode, certificateItems, maskedFC):
        self.logger.debug("REMOVE OLD DOCUMENT.")
        if waiverCode is None:
            waiverCode = ""

        if maskedFC is None:
            maskedFC = ""

        values = {
            'org.apache.struts.taglib.html.TOKEN': token,
            'amountCommission': commission,
            'miscSupportTypeId': "",
            'waiverCode': waiverCode,
            'certificateItem[0].value': certificateItems[0],
            'certificateItem[1].value': certificateItems[1],
            'certificateItem[2].value': certificateItems[2],
            'certificateItem[3].value': certificateItems[3],
            'error22010': "false",
            'oldDocumentAirlineCodeFI': "",
            'oldDocumentNumberFI': "",
            'removeOldDocumentButton.x': "13",
            'removeOldDocumentButton.y': "7",
            'selectedOldDocuments': "0",
            'newDocumentNumber': "",
            'maskedFC': maskedFC,
        }

        url = "https://iar2.arccorp.com/IAR/financialDetails.do"
        data = urllib.urlencode(values)

        headers = {
            'Accept': self._accept,
            'Accept-Encoding': self._accept_encoding,
            'Accept-Language': self._accept_language,
            'Connection': self._connection,
            'Content-Length': len(data),
            'Content-Type': self._content_type,
            'Host': self._host,
            'Referer': "https://iar2.arccorp.com/IAR/exchangeSummary.do",
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }

        req = urllib2.Request(url, data, headers)
        res = self.__try_request(req)
        if res:
            html = res.read()
            self.__save_page("remove_old_document", documentNumber, html)
            return html

    # def searchError(self, ped, action, arcNumber, token, from_date, to_date):
    #     self.logger.debug("SEARCH TICKET ERROR")
    #     values = {
    #         'org.apache.struts.taglib.html.TOKEN': token,
    #         'arcNumber': arcNumber,
    #         'ped': ped,
    #         'selectedStatusId': 'E',
    #         'documentNumber': "",
    #         'docNumberEnd': "",
    #         'selectedDocumentType': 'ET',
    #         'selectedTransactionType': '',
    #         'selectedFormOfPayment': '',
    #         'selectedInternationalIndicator': '',
    #         'systemProvider': '',
    #         'dateTypeRadioButtons': 'ped',
    #         'viewFromDate': from_date,
    #         'viewToDate': to_date,
    #         'commTypeRadioButtons': 'commEqualTo',
    #         'commissionAmount': '',
    #         'threeDigitCarrierCode': '',
    #         'selectedNumberOfResults': '500',
    #         'list.x': '45',
    #         'list.y': '11',
    #         'printOption': '1',
    #         'printaction': '0'
    #     }
    #
    #     url = "https://iar2.arccorp.com/IAR/listTransactions.do"
    #     data = urllib.urlencode(values)
    #
    #     headers = {
    #         'Accept': self._accept,
    #         'Accept-Encoding': 'gzip, deflate, br',
    #         'Accept-Language': self._accept_language,
    #         'Cache-Control': self._cache_control,
    #         'Connection': self._connection,
    #         'Content-Length': len(data),
    #         'Content-Type': self._content_type,
    #         'Host': 'iar2.arccorp.com',
    #         'Origin': 'https://iar2.arccorp.com',
    #         'Referer': 'https://iar2.arccorp.com/IAR/listTransactions.do?ped=' + ped + '&action=' + action + '&arcNumber=' + arcNumber + '',
    #         'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
    #         'User-Agent': self._user_agent
    #     }
    #     req = urllib2.Request(url, data, headers)
    #     res = self.__try_request(req)
    #     if res:
    #         html = res.read()
    #         self.__save_page("search", arcNumber, html)
    #         return html

    def search_error(self, ped, action, arcNumber, token, from_date, to_date, page_index, is_next=False):
        self.logger.debug("SEARCH TICKET ERROR, NEXT: %s" % is_next)
        values = {
            'org.apache.struts.taglib.html.TOKEN': token,
            'arcNumber': arcNumber,
            'ped': ped,
            'selectedStatusId': 'E',
            'documentNumber': "",
            'docNumberEnd': "",
            'selectedDocumentType': 'ET',
            'selectedTransactionType': '',
            'selectedFormOfPayment': '',
            'selectedInternationalIndicator': '',
            'systemProvider': '',
            'dateTypeRadioButtons': 'ped',
            'viewFromDate': from_date,
            'viewToDate': to_date,
            'commTypeRadioButtons': 'commEqualTo',
            'commissionAmount': '',
            'threeDigitCarrierCode': '',
            'selectedNumberOfResults': '500',
            'list.x': '45',
            'list.y': '11',
            'printOption': '1',
            'printaction': '0'
        }

        if is_next:
            del values["list.x"]
            del values["list.y"]

            values["next.x"] = "18"
            values["next.y"] = "10"

        url = "https://iar2.arccorp.com/IAR/listTransactions.do"
        data = urllib.urlencode(values)

        headers = {
            'Accept': self._accept,
            'Accept-Encoding': self._accept_encoding,
            'Accept-Language': self._accept_language,
            'Cache-Control': self._cache_control,
            'Connection': self._connection,
            'Content-Length': len(data),
            'Content-Type': self._content_type,
            'Host': self._host,
            'Origin': self._origin,
            'Referer': 'https://iar2.arccorp.com/IAR/listTransactions.do?ped=' + ped + '&action=' + action + '&arcNumber=' + arcNumber + '',
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }
        req = urllib2.Request(url, data, headers)
        res = self.__try_request(req)
        if res:
            html = res.read()
            self.__save_page("search" + str(page_index), arcNumber, html)
            return html

    def modifyTran(self, seqNum, documentNumber):
        self.logger.debug("GO TO MODIFY TRAN")
        url = "https://iar2.arccorp.com/IAR/modifyTran.do?"
        values = {
            'seqNum': seqNum,
            'documentNumber': documentNumber
        }
        data = urllib.urlencode(values)
        headers = {
            'Accept': self._accept,
            'Accept-Encoding': 'gzip, deflate, sdch, br',
            'Accept-Language': self._accept_language,
            'Connection': self._connection,
            'Host': self._host,
            'Referer': 'https://iar2.arccorp.com/IAR/listTransactions.do',
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }
        req = urllib2.Request(url, data, headers)
        res = self.__try_request(req)
        if res:
            html = res.read()
            self.__save_page("modifyTran", "modifyTran", html)
            return html

    def financialDetails(self, token, is_check_payment, commission, waiverCode, maskedFC, seqNum, documentNumber, tour_code,
                         qc_tour_code, certificates, certificate, agent_codes, is_et_button=False, is_check_update=False):
        self.logger.debug("GO TO FINANCIAL DETAILS")
        certificateItems = []
        if certificates:
            for i in certificates:
                if i[1] not in agent_codes:
                    certificateItems.append(i[1])

        certificateItem_len = len(certificateItems)
        for i in range(0, 3):
            if i >= certificateItem_len:
                certificateItems.append("")

        if is_check_payment:
            certificate = ""

        if not certificate and len(certificateItems) == 3:
            certificateItems.insert(3, "")
        elif certificate and len(certificateItems) == 3:
            certificateItems.insert(0, certificate)

        if not waiverCode:
            waiverCode = ""

        url = "https://iar2.arccorp.com/IAR/financialDetails.do"
        values = {
            'org.apache.struts.taglib.html.TOKEN': token,
            'navButton2.x': "63",
            'navButton2.y': "18",
            'amountCommission': commission,
            'miscSupportTypeId': "",
            'waiverCode': waiverCode,
            'certificateItem[0].value': certificateItems[0],
            'certificateItem[1].value': certificateItems[1],
            'certificateItem[2].value': certificateItems[2],
            'certificateItem[3].value': certificateItems[3],
            'error22010': "false",
            'oldDocumentAirlineCodeFI': "",
            'oldDocumentNumberFI': "",
            'maskedFC': maskedFC
            # 'ETButton.x':"27",
            # 'ETButton.y':"7"
        }

        if is_et_button or (not is_check_update and tour_code == qc_tour_code):
            del values['navButton2.x']
            del values['navButton2.y']
            values['ETButton.x'] = "27"
            values['ETButton.y'] = "7"

        data = urllib.urlencode(values)
        headers = {
            'Accept': self._accept,
            'Accept-Encoding': self._accept_encoding,
            'Accept-Language': self._accept_language,
            'Cache-Control': self._cache_control,
            'Connection': self._connection,
            'Content-Length': len(data),
            'Content-Type': self._content_type,
            'Host': self._host,
            'Origin': self._origin,
            'Referer': "https://iar2.arccorp.com/IAR/modifyTran.do?seqNum=" + seqNum + "&documentNumber=" + documentNumber,
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }

        req = urllib2.Request(url, data, headers)
        res = self.__try_request(req)
        if res:
            html = res.read()
            self.__save_page("FinancialDetails", "FinancialDetails", html)
            return html

    def financial_details(self, token, commission, waiverCode, certificateItems, maskedFC, is_et_button):
        self.logger.debug("GO TO FINANCIAL DETAILS")
        if waiverCode is None:
            waiverCode = ""

        if maskedFC is None:
            maskedFC = ""

        url = "https://iar2.arccorp.com/IAR/financialDetails.do"
        values = {
            'org.apache.struts.taglib.html.TOKEN': token,
            'navButton2.x': "63",
            'navButton2.y': "18",
            'amountCommission': commission,
            'miscSupportTypeId': "",
            'waiverCode': waiverCode,
            'certificateItem[0].value': certificateItems[0],
            'certificateItem[1].value': certificateItems[1],
            'certificateItem[2].value': certificateItems[2],
            'certificateItem[3].value': certificateItems[3],
            'error22010': "true",
            'oldDocumentAirlineCodeFI': "",
            'oldDocumentNumberFI': "",
            'maskedFC': maskedFC
            # 'ETButton.x': "47",
            # 'ETButton.y': "7"
        }

        if is_et_button:
            del values['navButton2.x']
            del values['navButton2.y']
            values['ETButton.x'] = "27"
            values['ETButton.y'] = "7"

        data = urllib.urlencode(values)
        headers = {
            'Accept': self._accept,
            'Accept-Encoding': self._accept_encoding,
            'Accept-Language': self._accept_language,
            'Connection': self._connection,
            'Content-Length': len(data),
            'Content-Type': self._content_type,
            'Host': self._host,
            'Referer': "https://iar2.arccorp.com/IAR/financialDetails.do",
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }

        req = urllib2.Request(url, data, headers)
        res = self.__try_request(req)
        if res:
            html = res.read()
            self.__save_page("FinancialDetails", "FinancialDetails", html)
            return html

    def itineraryEndorsements(self, token, qc_tour_code, backOfficeRemarks, ticketDesignators):
        self.logger.debug("GO TO ITINERARY ENDORSEMENTS")
        url = "https://iar2.arccorp.com/IAR/itineraryEndorsements.do"
        if not backOfficeRemarks:
            backOfficeRemarks = ""
        values = {
            'org.apache.struts.taglib.html.TOKEN': token,
            'tourCode': qc_tour_code,
            # 'coupon[0].ticketDesignator':"",
            # 'coupon[1].ticketDesignator':"",
            # 'coupon[2].ticketDesignator':"",
            # 'coupon[3].ticketDesignator':"",
            'backOfficeRemarks': backOfficeRemarks,
            'ETButton.x': "32",
            'ETButton.y': "8"
        }
        if ticketDesignators:
            for i in ticketDesignators:
                values['coupon[' + i[0] + '].ticketDesignator'] = i[1]
        else:
            values['coupon[0].ticketDesignator'] = ""
            values['coupon[1].ticketDesignator'] = ""
            values['coupon[2].ticketDesignator'] = ""
            values['coupon[3].ticketDesignator'] = ""

        data = urllib.urlencode(values)
        headers = {
            'Accept': self._accept,
            'Accept-Encoding': self._accept_encoding,
            'Accept-Language': self._accept_language,
            'Cache-Control': self._cache_control,
            'Connection': self._connection,
            'Content-Length': len(data),
            'Content-Type': self._content_type,
            'Host': self._host,
            'Origin': self._origin,
            'Referer': "https://iar2.arccorp.com/IAR/financialDetails.do",
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }

        req = urllib2.Request(url, data, headers)
        res = self.__try_request(req)
        if res:
            html = res.read()
            self.__save_page("ItineraryEndorsements", "ItineraryEndorsements", html)
            return html

    def transactionConfirmation(self, token):
        self.logger.debug("TRANSACTION CONFIRMATION")
        url = "https://iar2.arccorp.com/IAR/transactionConfirmation.do"
        values = {
            'org.apache.struts.taglib.html.TOKEN': token,
            'selected': "201",
            'yes.x': "24",
            'yes.y': "4"
        }
        data = urllib.urlencode(values)
        headers = {
            'Accept': self._accept,
            'Accept-Encoding': self._accept_encoding,
            'Accept-Language': self._accept_language,
            'Cache-Control': self._cache_control,
            'Connection': self._connection,
            'Content-Length': len(data),
            'Content-Type': self._content_type,
            'Host': self._host,
            'Origin': self._origin,
            'Referer': "https://iar2.arccorp.com/IAR/itineraryEndorsements.do",
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }

        req = urllib2.Request(url, data, headers)
        res = self.__try_request(req)
        if res:
            html = res.read()
            self.__save_page("TransactionConfirmation", "TransactionConfirmation", html)
            return html

    def iar_logout(self, ped, action, arcNumber):
        # print "iar logout Start"
        self.logger.debug("IAR LOGOUT START")
        url = "https://iar2.arccorp.com/IAR/logout.do"
        headers = {
            'Accept': self._accept,
            'Accept-Encoding': 'gzip, deflate, sdch',
            'Accept-Language': self._accept_language,
            'Connection': self._connection,
            'Host': self._host,
            'Referer': 'https://iar2.arccorp.com/IAR/listTransactions.do?ped=' + ped + '&action=' + action + '&arcNumber=' + arcNumber,
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }
        req = urllib2.Request(url, None, headers)
        self.__try_request(req)
        self.logger.debug("Iar logout end")

    def logout(self):
        # print 'Logout Start'
        self.logger.debug('LOGOUT START')
        headers = {
            'Accept': self._accept,
            'Accept-Encoding': 'gzip, deflate, sdch, br',
            'Accept-Language': self._accept_language,
            'Connection': self._connection,
            'Host': 'myarc.arccorp.com',
            # 'Referer':'https://iar2.arccorp.com/IAR/listTransactions.do?ped='+ped+'&action='+action+'&arcNumber=45668571',
            'Referer': 'https://myarc.arccorp.com/PortalApp/ARCGateway.portal?_nfpb=true&_st=&_pageLabel=ARC_Home&_nfls=false',
            'Upgrade-Insecure-Requests': self._upgrade_insecure_requests,
            'User-Agent': self._user_agent
        }
        url = "https://myarc.arccorp.com/PortalApp/ARCGateway.portal?_nfpb=true&_st=&_pageLabel=ARC_Login&_name=logout&_nfls=false"
        req = urllib2.Request(url, None, headers)
        self.__try_request(req)
        self.logger.debug("Logout end")

    def store(self, data, arc_number=""):
        path = 'file'
        if not os.path.exists(path):
            os.makedirs(path)
        today = datetime.datetime.now().strftime('%Y%m%d')
        if arc_number:
            today = today + "_" + arc_number
        file_name = today + '.json'
        with open(path + '\\' + file_name, 'w') as json_file:
            json_file.write(json.dumps(data))

    def load(self):
        today = datetime.datetime.now().strftime('%Y%m%d')
        file_name = 'file\\' + today + '.json'
        if not os.path.isfile(file_name):
            return None
        with open(file_name) as json_file:
            data = json.load(json_file)
            return data

    def read_file(self, path, file_name):
        file_full_path = path + '/' + file_name
        if not os.path.isfile(file_full_path):
            return None

        with open(file_full_path, 'r') as f:
            return f.read()

    def convertStatus(self, data, is_remove_error=False):
        status = 'No'
        is_updated = 'Fail'

        if data['Status'] == 1:
            status = 'Yes'
        elif data['Status'] == 2:
            status = 'Void'
        elif data['Status'] == 3:
            status = 'CK'
        elif data['Status'] == 4:
            status = 'Pass'

        if is_remove_error:
            is_updated = ""
            return status, is_updated

        if data['Status'] == 1 or data['Status'] == 3:
            if "isPutError" not in data or not data['isPutError']:
                data['QCComm'] = "" if data['QCComm'] is None else data['QCComm']
                data['ArcCommUpdated'] = "" if data['ArcCommUpdated'] is None else data['ArcCommUpdated']
                data['QCTourCode'] = "" if data['QCTourCode'] is None else data['QCTourCode'].upper()
                data['ArcTourCodeUpdated'] = "" if data['ArcTourCodeUpdated'] is None else data['ArcTourCodeUpdated']
                if data['QCComm'] == data['ArcCommUpdated'] and data['QCTourCode'] == data['ArcTourCodeUpdated']:
                    is_updated = 'Success'
            else:
                if data['hasPutError']:
                    is_updated = 'Success'
        elif data['Status'] == 2:
            is_updated = 'Void'
        elif data['Status'] == 4:
            is_updated = 'Pass'

        return status, is_updated

    def exportExcel(self, datas, file_name):
        path = 'excel'
        if not os.path.exists(path):
            os.makedirs(path)
        wb = Workbook()
        ws = wb.active
        ws.column_dimensions['K'].width = 0
        ws['A1'] = 'ARC'
        ws['B1'] = 'Ticket number'
        ws['C1'] = 'Date'
        ws['D1'] = 'TKCM'
        ws['E1'] = 'QCCM'
        ws['F1'] = 'ARCCM'
        ws['G1'] = 'TKTC'
        ws['H1'] = 'QCTC'
        ws['I1'] = 'ARCTC'
        ws['J1'] = 'Status'
        ws['K1'] = 'Result'
        row_index = 2
        for data in datas:
            ws.cell(row=row_index, column=1).value = data['ArcNumber']
            ws.cell(row=row_index, column=2).value = data['TicketNumber']
            ws.cell(row=row_index, column=3).value = data['IssueDate']
            ws.cell(row=row_index, column=4).value = data['Comm']
            ws.cell(row=row_index, column=5).value = data['QCComm']
            ws.cell(row=row_index, column=6).value = data['ArcCommUpdated']
            ws.cell(row=row_index, column=7).value = data['TourCode']
            ws.cell(row=row_index, column=8).value = data['QCTourCode']
            ws.cell(row=row_index, column=9).value = data['ArcTourCodeUpdated']

            status, is_updated = self.convertStatus(data)

            ws.cell(row=row_index, column=10).value = is_updated
            ws.cell(row=row_index, column=11).value = status
            row_index += 1
        wb.save(filename=path + "/" + file_name + ".xlsx")


class Regex:
    def __init__(self):
        self._pattern_token = re.compile(r'input type="hidden" name="org\.apache\.struts\.taglib\.html\.TOKEN" value="([\da-z]{32})"')
        self._pattern_search = re.compile(r'<a href="/IAR/modifyTran\.do\?seqNum=(\d{10})&amp;documentNumber=(\d{10})"')
        self._pattern_masked = re.compile(
            r'<textarea name="maskedFC" cols="60" rows="5" readonly="readonly" class="disabled">(.+?)</textarea>')
        self._pattern_commission = re.compile(
            r'<input type="text" name="amountCommission" maxlength="10" size="15" value="(\d+\.\d{2}|\s?)" (disabled="disabled" )?class="contenttextright">')
        self._pattern_waiver_code = re.compile(
            r'<input name="waiverCode" maxlength="15" size="20" value="(.+?)" class="contenttext" type="text">')
        self._pattern_certificates = re.compile(
            r'<input type="text" name="certificateItem\[(\d{1})\]\.value" maxlength="14" size="19" value="(.+?)" class="contenttext">')
        self._pattern_tour_code = re.compile(
            r'<input type="text" name="tourCode" maxlength="15" size="22" value="(.*?)" (disabled="disabled" )?class="contenttext">')
        self._pattern_backOfficeRemarks = re.compile(
            r'<input type="text" name="backOfficeRemarks" maxlength="49" size="70" value="(.*?)" (disabled="disabled" )?class="contenttext">')
        self._pattern_ticketDesignator = re.compile(
            r'<input type="text" name="coupon\[(\d{1})\]\.ticketDesignator" maxlength="14" size="20" value="(.*?)" (disabled="disabled" )?class="contenttext">')
        # self._pattern_tran_type = re.compile(r'''                      <td width="30%" height="24" class="contentboldtext" align="left">&nbsp;Tran Type:</td>
        #               <td width="35%" class="contenttext">(.*?)</td>
        #               <td width="20%" class="contentboldtext"  align="left">Status:&nbsp;</td>
        #               <td width="14%">(.*?)</td>''')
        self._pattern_tran_type = re.compile(r'''                      <td width="30%" height="24" class="contentboldtext" align="left">&nbsp;Tran Type:</td>
                      <td width="35%" class="contenttext">(.*?)</td>''')
        self._pattern_modify_trans = re.compile(r'''<td width="7%" align="center">(\d{3})</td>
        <td width="11%" align="left">
        
        
			<a href="/IAR/modifyTran\.do\?seqNum=(\d{10})&amp;documentNumber=(\d{10})".+?
            
                
        
		</td>
        <td width="4%" align="right" >(.*?) 
        </td>''')
        self._pattern_total = re.compile(r'<input type="text" name="amountTotal" maxlength="12" size="15" value="(.*?)"')
        # pass

    def __public(self, pattern, html, is_findall=False):
        if is_findall:
            return pattern.findall(html)
        else:
            m = pattern.search(html)
            if m:
                return m.groups()

    def __token(self, html):
        result = self.__public(self._pattern_token, html)
        if result:
            return result[0]

    def iar(self, html, is_this_week=True):
        ped = action = arcNumber = None
        date_time = datetime.datetime.now()
        # date_week = date_time.weekday()
        date_ped = date_time + datetime.timedelta(days=(6 - date_time.weekday()))

        if not is_this_week:
            date_ped = date_ped + datetime.timedelta(days=-7)

        ped = date_ped.strftime('%d%b%y').upper()
        pattern = re.compile(r'<a href="/IAR/listTransactions\.do;.+?ped=' + ped + '&amp;action=(\d)&amp;arcNumber=(\d{8})">')
        result = self.__public(pattern, html)
        if result:
            action = result[0]
            arcNumber = result[1]
        return ped, action, arcNumber

    def create_list(self, html):
        return self.__token(html)

    def get_token(self, html):
        return self.__token(html)

    def listTransactions(self, html):
        token = from_date = to_date = None

        token_index = html.find("name=\"org.apache.struts.taglib.html.TOKEN\" value=\"")
        if token_index >= 0:
            token_sub = html[token_index:]
            token_value_end = token_sub.find("\">")
            token_value = token_sub[:token_value_end]
            token = token_value[-32:]

        from_date_index = html.find("name=\"viewFromDate\"")
        if from_date_index >= 0:
            from_date_value_index = html[from_date_index:].find("value=\"")
            from_date = html[from_date_index + from_date_value_index + 7:from_date_index + from_date_value_index + 14]

        to_date_index = html.find("name=\"viewToDate\"")
        if to_date_index >= 0:
            to_date_value_index = html[to_date_index:].find("value=\"")
            to_date = html[to_date_index + to_date_value_index + 7:to_date_index + to_date_value_index + 14]
        return token, from_date, to_date

    def search(self, html):
        result = self.__public(self._pattern_search, html)
        if result:
            return result
        else:
            return None, None

    def modifyTran(self, html):
        token = maskedFC = waiverCode = commission = None
        token = self.__token(html)

        masked_result = self.__public(self._pattern_masked, html)
        if masked_result:
            maskedFC = masked_result[0]

        commission_result = self.__public(self._pattern_commission, html)
        if commission_result:
            commission = commission_result[0]

        waiverCode_result = self.__public(self._pattern_waiver_code, html)
        if waiverCode_result:
            waiverCode = waiverCode_result[0]

        certificates_result = self.__public(self._pattern_certificates, html, True)

        return token, maskedFC, commission, waiverCode, certificates_result

    def financialDetails(self, html):
        token = tour_code = backOfficeRemarks = None

        token = self.__token(html)

        tour_code_result = self.__public(self._pattern_tour_code, html)
        if tour_code_result:
            tour_code = tour_code_result[0]

        backOfficeRemarks_result = self.__public(self._pattern_backOfficeRemarks, html)
        if backOfficeRemarks_result:
            backOfficeRemarks = backOfficeRemarks_result[0]

        ticketDesignator_result = self.__public(self._pattern_ticketDesignator, html, True)
        return token, tour_code, backOfficeRemarks, ticketDesignator_result

    def itineraryEndorsements(self, html):
        return self.__token(html)

    def search_error(self, html, entry_date, error_codes):
        pattern = re.compile(r'''<tr align=right class="row[01]"> 
          <td width="3%"  align="center"> .+? 
          </td>
          <td width="6%" align="center">
				 <!-- defect #810-->
				 <input type="checkbox" name="checkBoxes" value='(\d{10}):(\d{10})'  />  
          <td  width="5%" align="center">E 
          </td>
          <td width="6%"align="center">(\d{3}) 
          </td>
        <td width="12%" align="left">
        
            <a href="/IAR/modifyTran\.do\?seqNum=.+?</a>
                
        
		</td>
          <td width="14%" align="left">.+?      
          </td>         
          <td width="21%" align="left">.+? 
          </td>
          <td width="16%" align="left">(''' + error_codes + ''') 
          </td>
          <td width="10%"  align="center">(''' + entry_date + ''') 
          </td>
          <td width="10%"  align="center" nowrap>.+? 
          </td>
        </tr>''')
        return self.__public(pattern, html, True)

    def tran_type(self, html):
        result = self.__public(self._pattern_tran_type, html)
        if result:
            return result[0]

    def modify_trans(self, html):
        return self.__public(self._pattern_modify_trans, html, True)

    def get_total(self, html):
        return self.__public(self._pattern_total, html)

    # 1 = pass, 2 = void, 0 = default
    def check_status(self, html):
        status = 0
        if html.find('Document is being displayed as view only') >= 0:
            status = 1
            if html.find('Unable to modify a voided document') >= 0:
                status = 2

        return status


class MSSQL:
    def __init__(self, server, db, user, pwd):
        self.server = server
        self.db = db
        self.user = user
        self.pwd = pwd

    def __GetConnect(self):
        if not self.db:
            raise (NameError, "database no value")
        self.conn = pyodbc.connect(DRIVER="{SQL Server}", SERVER=self.server, DATABASE=self.db, UID=self.user,
                                   PWD=self.pwd)

        cur = self.conn.cursor()
        if not cur:
            raise (NameError, "connnect error")
        else:
            return cur

    def ExecQuery(self, sql):
        cur = self.__GetConnect()
        cur.execute(sql)
        resList = cur.fetchall()

        self.conn.close()
        return resList

    def ExecNonQuery(self, sql):
        cur = self.__GetConnect()
        rowcount = cur.execute(sql).rowcount
        self.conn.commit()
        self.conn.close()
        return rowcount

    def ExecNonQuerys(self, sqls):
        cur = self.__GetConnect()
        rowcount = 0
        for sql in sqls:
            rowcount += cur.execute(sql).rowcount
        self.conn.commit()
        self.conn.close()
        return rowcount

    # def ExecNonQueryMany(self, values):
    #     cur = self.__GetConnect()
    #     cur.executemany("insert into test(id,name,age) values (?,?,?);", values)
    #     self.conn.commit()
    #     self.conn.close()


class Email:
    """docstring for SendEmail"""

    def __init__(self, is_local=True, smtp_server=None, user=None, password=None):
        # self.from_addr=from_addr
        # self.to_addr=to_addr
        self.smtp_server = smtp_server
        self.user = user
        self.password = password
        # self.subject=subject
        self.is_local = is_local

    def __format_addr(self, s):
        name, addr = parseaddr(s)
        return formataddr(( \
            Header(name, 'utf-8').encode(), \
            addr.encode('utf-8') if isinstance(addr, unicode) else addr))

    def send(self, from_addr, to_addr, subject, body, files=[]):
        # msg = MIMEText(body, 'plain', 'utf-8')
        # msg = MIMEText(body, 'html', 'utf-8')
        msg = MIMEMultipart()
        # msg['From'] = self.from_addr #self._format_addr('post<%s>' % from_addr)
        msg['From'] = self.__format_addr('no-reply<%s>' % from_addr)
        msg['To'] = ";".join(to_addr)
        msg['Subject'] = Header(subject, 'utf-8').encode()
        msg.attach(MIMEText(body, 'html', 'utf-8'))
        for f in files:
            part = MIMEBase('application', "octet-stream")
            part.set_payload(open(f, "rb").read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(f))
            msg.attach(part)
        server = smtplib.SMTP()
        # server.set_debuglevel(1)
        try:
            server.connect(self.smtp_server, 25)
            # server = smtplib.SMTP(self.smtp_server, 25)

            if not self.is_local:
                # print 'remote'
                server.login(self.user, self.password)
            # else:
            # 	print 'local'
            server.sendmail(from_addr, to_addr, msg.as_string())
        # print 'sent'
        except Exception as e:
            print e
        finally:
            server.quit()


# class SendEmail:
#     """docstring for SendEmail"""
#
#     def __init__(self, from_addr, to_addr, smtp_server, subject):
#         self.from_addr = from_addr
#         self.to_addr = to_addr
#         self.smtp_server = smtp_server
#         self.subject = subject
#
#     def __format_addr(self, s):
#         name, addr = parseaddr(s)
#         return formataddr(( \
#             Header(name, 'utf-8').encode(), \
#             addr.encode('utf-8') if isinstance(addr, unicode) else addr))
#
#     def send(self, body):
#         # msg = MIMEText(body, 'plain', 'utf-8')
#         msg = MIMEText(body, 'html', 'utf-8')
#         # msg['From'] = self.from_addr #self._format_addr('post<%s>' % from_addr)
#         msg['From'] = self.__format_addr('no-reply<%s>' % self.from_addr)
#         msg['To'] = ";".join(self.to_addr)
#         msg['Subject'] = Header(self.subject, 'utf-8').encode()
#
#         server = smtplib.SMTP(self.smtp_server, 25)
#         # server.set_debuglevel(1)
#         try:
#             # if not is_local:
#             # 	server.login(from_addr, password)
#             # else:
#             # 	print 'local'
#             server.sendmail(self.from_addr, self.to_addr, msg.as_string())
#         # print 'sent'
#         except Exception as e:
#             print e
#         finally:
#             server.quit()